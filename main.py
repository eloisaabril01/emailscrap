
from flask import Flask, request, render_template, jsonify
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import re, time, requests, subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import hashlib

app = Flask(__name__)

# JSON database file
DB_FILE = "database.json"

def load_db():
    """Load database from JSON file"""
    if os.path.exists(DB_FILE):
        with open(DB_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_db(data):
    """Save database to JSON file"""
    with open(DB_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def get_excel_filename(query):
    """Generate Excel filename from query"""
    # Clean query to make it safe for filename
    safe_query = re.sub(r'[^\w\s-]', '', query).strip().replace(' ', '_')
    return f"exports/{safe_query}.xlsx"

def save_to_excel(query, results):
    """Save or append results to Excel file for the given query"""
    if not results:
        return
    
    # Create exports directory if it doesn't exist
    os.makedirs("exports", exist_ok=True)
    
    filename = get_excel_filename(query)
    
    # Check if file exists
    if os.path.exists(filename):
        # Load existing workbook
        wb = load_workbook(filename)
        ws = wb.active
        # Find the next available row
        next_row = ws.max_row + 1
        # Get the current max sr_no
        last_sr_no = ws.cell(ws.max_row, 1).value if ws.max_row > 1 else 0
        if not isinstance(last_sr_no, int):
            last_sr_no = 0
    else:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Create header row with styling
        headers = ["Sr No", "Name", "Address", "Phone", "Website", "Emails", "Date Added"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 20
        
        next_row = 2
        last_sr_no = 0
    
    # Add new results
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    for result in results:
        sr_no = last_sr_no + 1
        last_sr_no = sr_no
        
        # Create business identifier for duplicate checking within Excel
        business_id = f"{result['name']}|{result['address']}"
        
        # Check if this business already exists in the Excel file
        already_exists = False
        if next_row > 2:  # If there are existing rows beyond header
            for row in range(2, next_row):
                existing_name = ws.cell(row, 2).value
                existing_address = ws.cell(row, 3).value
                if existing_name == result['name'] and existing_address == result['address']:
                    already_exists = True
                    break
        
        if already_exists:
            continue
        
        ws.cell(next_row, 1, sr_no)
        ws.cell(next_row, 2, result['name'])
        ws.cell(next_row, 3, result['address'])
        ws.cell(next_row, 4, result['phone'])
        ws.cell(next_row, 5, result['website'] if result['website'] else "N/A")
        ws.cell(next_row, 6, ", ".join(result['emails']) if result['emails'] else "N/A")
        ws.cell(next_row, 7, current_date)
        
        # Apply alignment
        for col in range(1, 8):
            ws.cell(next_row, col).alignment = Alignment(vertical="top", wrap_text=True)
        
        next_row += 1
    
    # Save the workbook
    wb.save(filename)
    return filename

# Global variable to track progress
progress_data = {"current": 0, "total": 0, "status": "idle"}
progress_lock = threading.Lock()

# Global stop flag
stop_scraping = False
stop_lock = threading.Lock()

def get_chrome_driver():
    """Create and return a configured Chrome driver instance"""
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--single-process")  # Reduce process overhead
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-background-networking")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    # Add experimental options to avoid detection
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    try:
        chromium_path = subprocess.check_output(['which', 'chromium']).decode().strip()
        chromedriver_path = subprocess.check_output(['which', 'chromedriver']).decode().strip()
        options.binary_location = chromium_path
        service = Service(executable_path=chromedriver_path)
        driver = webdriver.Chrome(service=service, options=options)
    except Exception as e:
        print(f"Error creating Chrome driver with custom paths: {e}")
        driver = webdriver.Chrome(options=options)
    
    # Set page load timeout
    driver.set_page_load_timeout(20)
    return driver

def is_business_already_shown(business_id, query):
    """Check if a business has already been shown for this query"""
    db = load_db()
    db_key = f"shown_{query}"
    if db_key in db:
        shown_businesses = db[db_key]
        return business_id in shown_businesses
    return False

def mark_business_as_shown(business_id, query):
    """Mark a business as shown for this query"""
    db = load_db()
    db_key = f"shown_{query}"
    if db_key in db:
        shown_businesses = db[db_key]
        if business_id not in shown_businesses:
            shown_businesses.append(business_id)
            db[db_key] = shown_businesses
    else:
        db[db_key] = [business_id]
    save_db(db)

def fetch_emails_from_website(website):
    """Fetch emails from a website with retry limit"""
    if not website:
        return []
    
    max_attempts = 2
    for attempt in range(max_attempts):
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            r = requests.get(website, timeout=3, headers=headers, allow_redirects=True)
            found_emails = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", r.text)
            
            # Filter out common fake/example emails and invalid patterns
            invalid_patterns = ['example.com', 'test.com', 'wixpress.com', 'sentry.io', 'placeholder', 
                              'yourdomain', 'domain.com', '.jpg', '.png', '.gif', '.jpeg', '.svg', 
                              '@2x', 'image', 'photo', 'picture']
            
            emails = [e for e in set(found_emails) 
                     if not any(x in e.lower() for x in invalid_patterns)
                     and '@' in e 
                     and '.' in e.split('@')[1]][:3]
            return emails
        except Exception as e:
            if attempt < max_attempts - 1:
                time.sleep(0.3)  # Quick retry
            else:
                print(f"⚠️ Skipping {website[:40]} after {max_attempts} attempts")
                return []
    return []

def verify_email(email):
    """Basic email verification (format check)"""
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(email_pattern, email) is not None

def process_single_card(driver, card, idx):
    """Process a single Google Maps card"""
    try:
        driver.execute_script("arguments[0].scrollIntoView(true);", card)
        time.sleep(0.3)
        card.click()
        time.sleep(1)

        name = "N/A"
        try:
            name_elem = driver.find_element(By.CSS_SELECTOR, "h1.DUwDvf")
            name = name_elem.text
        except:
            pass

        address = "N/A"
        try:
            address_elem = driver.find_element(By.CSS_SELECTOR, "button[data-item-id*='address']")
            address = address_elem.text
        except:
            try:
                address_elem = driver.find_element(By.CSS_SELECTOR, "div[data-item-id*='address'] div.fontBodyMedium")
                address = address_elem.text
            except:
                pass

        phone = "N/A"
        try:
            phone_elem = driver.find_element(By.CSS_SELECTOR, "button[data-item-id*='phone']")
            phone = phone_elem.text
        except:
            try:
                phone_elem = driver.find_element(By.CSS_SELECTOR, "div[data-item-id*='phone'] div.fontBodyMedium")
                phone = phone_elem.text
            except:
                pass

        website = None
        try:
            website_elem = driver.find_element(By.CSS_SELECTOR, "a[data-item-id='authority']")
            website = website_elem.get_attribute("href")
        except:
            try:
                website_buttons = driver.find_elements(By.CSS_SELECTOR, "a[data-tooltip='Open website']")
                if website_buttons:
                    website = website_buttons[0].get_attribute("href")
            except:
                pass
        
        if website and "google.com" in website:
            website = None

        return {
            "name": name,
            "address": address,
            "phone": phone,
            "website": website if website else None,
            "emails": []
        }
    except Exception as e:
        print(f"Error processing card {idx}: {e}")
        return None

def scrape_google_maps(query, limit=10):
    global progress_data, stop_scraping
    
    with stop_lock:
        stop_scraping = False
    
    with progress_lock:
        progress_data = {"current": 0, "total": limit, "status": "Initializing..."}
    
    seen_businesses = set()
    results_with_emails = []
    
    driver = get_chrome_driver()
    
    try:
        with progress_lock:
            progress_data["status"] = "Loading Google Maps..."
        
        driver.get(f"https://www.google.com/maps/search/{query.replace(' ', '+')}/")
        time.sleep(3)
        
        scrollable_div = driver.find_element(By.CSS_SELECTOR, "div[role='feed']")
        
        BATCH_SIZE = 5
        scroll_count = 0
        max_scrolls = 200
        consecutive_empty_batches = 0
        max_empty_batches = 5  # Stop if we get 5 consecutive empty batches
        
        while len(results_with_emails) < limit and scroll_count < max_scrolls and consecutive_empty_batches < max_empty_batches:
            with stop_lock:
                if stop_scraping:
                    break
            
            # Scroll to get more cards - use JavaScript scroll with fallback
            retry_count = 0
            max_retries = 2
            scroll_success = False
            
            while retry_count < max_retries and not scroll_success:
                try:
                    # Try to find and scroll the feed element
                    scrollable_div = driver.find_element(By.CSS_SELECTOR, "div[role='feed']")
                    driver.execute_script('arguments[0].scrollTop = arguments[0].scrollHeight', scrollable_div)
                    scroll_success = True
                    time.sleep(0.5)
                except Exception as e:
                    retry_count += 1
                    if retry_count >= max_retries:
                        print(f"Skipping scroll after {max_retries} attempts")
                        # Try alternative scroll method once
                        try:
                            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                            scroll_success = True
                        except:
                            print("Scroll failed, moving on...")
                            break
                    else:
                        time.sleep(0.5)
            
            scroll_count += 1
            
            # Get all visible cards
            cards = driver.find_elements(By.CSS_SELECTOR, "a.hfpxzc")
            
            # Collect batch of card URLs
            batch_urls = []
            for card in cards:
                if len(batch_urls) >= BATCH_SIZE:
                    break
                card_url = card.get_attribute("href")
                if card_url and card_url not in seen_businesses:
                    batch_urls.append(card_url)
            
            if not batch_urls:
                consecutive_empty_batches += 1
                print(f"No new cards found ({consecutive_empty_batches}/{max_empty_batches} empty batches)")
                if consecutive_empty_batches >= max_empty_batches:
                    print("Too many empty batches, stopping search...")
                    break
                continue
            else:
                consecutive_empty_batches = 0  # Reset counter when we find new cards
            
            with progress_lock:
                progress_data["status"] = f"Processing batch of {len(batch_urls)} businesses..."
            
            # Process batch: extract website info
            batch_with_websites = []
            for idx, card_url in enumerate(batch_urls):
                if len(results_with_emails) >= limit:
                    break
                
                with stop_lock:
                    if stop_scraping:
                        break
                
                try:
                    driver.get(card_url)
                    time.sleep(0.8)
                    
                    name = "N/A"
                    try:
                        name_elem = driver.find_element(By.CSS_SELECTOR, "h1.DUwDvf")
                        name = name_elem.text
                    except:
                        pass
                    
                    address = "N/A"
                    try:
                        address_elem = driver.find_element(By.CSS_SELECTOR, "button[data-item-id*='address']")
                        address = address_elem.text
                    except:
                        try:
                            address_elem = driver.find_element(By.CSS_SELECTOR, "div[data-item-id*='address'] div.fontBodyMedium")
                            address = address_elem.text
                        except:
                            pass
                    
                    phone = "N/A"
                    try:
                        phone_elem = driver.find_element(By.CSS_SELECTOR, "button[data-item-id*='phone']")
                        phone = phone_elem.text
                    except:
                        try:
                            phone_elem = driver.find_element(By.CSS_SELECTOR, "div[data-item-id*='phone'] div.fontBodyMedium")
                            phone = phone_elem.text
                        except:
                            pass
                    
                    website = None
                    try:
                        website_elem = driver.find_element(By.CSS_SELECTOR, "a[data-item-id='authority']")
                        website = website_elem.get_attribute("href")
                    except:
                        try:
                            website_buttons = driver.find_elements(By.CSS_SELECTOR, "a[data-tooltip='Open website']")
                            if website_buttons:
                                website = website_buttons[0].get_attribute("href")
                        except:
                            pass
                    
                    if website and "google.com" not in website:
                        business_id = f"{name}|{address}"
                        
                        if business_id not in seen_businesses and not is_business_already_shown(business_id, query):
                            seen_businesses.add(business_id)
                            batch_with_websites.append({
                                "name": name,
                                "address": address,
                                "phone": phone,
                                "website": website,
                                "business_id": business_id
                            })
                            print(f"✓ Card {idx+1}/{len(batch_urls)}: {name[:30]}... has website")
                except Exception as e:
                    print(f"Error processing card: {e}")
            
            # Extract emails from batch websites
            with progress_lock:
                progress_data["status"] = f"Extracting emails from {len(batch_with_websites)} websites..."
            
            def extract_and_verify(business_data):
                """Extract and verify emails for a business"""
                emails = fetch_emails_from_website(business_data['website'])
                if emails:
                    print(f"✉️  Found {len(emails)} emails from {business_data['name'][:30]}: {emails}")
                    verified = []
                    for email in emails:
                        if verify_email(email):
                            verified.append(email)
                            print(f"✅ VERIFIED: {email}")
                        else:
                            print(f"❌ Invalid: {email}")
                    return (business_data, verified)
                else:
                    print(f"❌ No emails found on {business_data['website'][:50]}")
                return (business_data, [])
            
            # Process websites in parallel
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = [executor.submit(extract_and_verify, biz) for biz in batch_with_websites]
                
                for future in as_completed(futures):
                    if len(results_with_emails) >= limit:
                        break
                    
                    with stop_lock:
                        if stop_scraping:
                            break
                    
                    try:
                        business_data, verified_emails = future.result()
                        
                        # If this business has verified emails, add it as ONE result with ALL emails
                        if verified_emails:
                            if len(results_with_emails) >= limit:
                                break
                            
                            result = {
                                "name": business_data['name'],
                                "address": business_data['address'],
                                "phone": business_data['phone'],
                                "website": business_data['website'],
                                "emails": verified_emails  # All emails from this domain
                            }
                            
                            result_hash = hashlib.md5(business_data['business_id'].encode()).hexdigest()
                            if not any(r.get('_hash') == result_hash for r in results_with_emails):
                                result['_hash'] = result_hash
                                results_with_emails.append(result)
                                mark_business_as_shown(business_data['business_id'], query)
                                
                                with current_results_lock:
                                    current_results.append(result)
                                
                                with progress_lock:
                                    progress_data["current"] = len(results_with_emails)
                                    progress_data["status"] = f"✓ {len(results_with_emails)}/{limit} businesses with verified emails"
                    except Exception as e:
                        print(f"Error in email extraction: {e}")
            
            # If we got enough results, stop
            if len(results_with_emails) >= limit:
                break
            
            with progress_lock:
                progress_data["status"] = f"Found {len(results_with_emails)}/{limit}, getting next batch..."
    
    finally:
        driver.quit()
    
    # Remove hash field before saving
    clean_results = []
    for r in results_with_emails:
        clean_r = {k: v for k, v in r.items() if k != '_hash'}
        clean_results.append(clean_r)
    
    # Save results
    if clean_results:
        with progress_lock:
            progress_data["status"] = "Saving to Excel..."
        try:
            excel_file = save_to_excel(query, clean_results)
            with progress_lock:
                progress_data["status"] = f"✓ Complete! Saved {len(clean_results)} results to {os.path.basename(excel_file)}"
        except Exception as e:
            with progress_lock:
                progress_data["status"] = f"✓ Complete! Found {len(clean_results)} results"
    else:
        with progress_lock:
            progress_data["status"] = f"✓ Complete! Found {len(clean_results)} verified results"
    
    return clean_results

# Global variable to store current results
current_results = []
current_results_lock = threading.Lock()

@app.route("/progress")
def get_progress():
    """Return current progress as JSON"""
    with progress_lock:
        return jsonify(progress_data)

@app.route("/stop", methods=["POST"])
def stop_search():
    """Stop the current search"""
    global stop_scraping
    with stop_lock:
        stop_scraping = True
    return jsonify({"status": "stopping"})

@app.route("/results")
def get_results():
    """Return current results as JSON"""
    with current_results_lock:
        return jsonify(current_results)

@app.route("/excel_files")
def list_excel_files():
    """Return list of available Excel files"""
    if not os.path.exists("exports"):
        return jsonify([])
    
    files = []
    for filename in os.listdir("exports"):
        if filename.endswith(".xlsx"):
            filepath = os.path.join("exports", filename)
            stat = os.stat(filepath)
            files.append({
                "name": filename,
                "size": f"{stat.st_size / 1024:.2f} KB",
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            })
    
    return jsonify(sorted(files, key=lambda x: x['modified'], reverse=True))

@app.route("/download/<filename>")
def download_file(filename):
    """Download an Excel file"""
    from flask import send_file
    filepath = os.path.join("exports", filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return "File not found", 404

@app.route("/view_excel")
def view_excel():
    """View all Excel files in a separate page"""
    return render_template("view_excel.html")

@app.route("/create_combined_excel")
def create_combined_excel():
    """Create a combined Excel file with all data"""
    if not os.path.exists("exports"):
        return jsonify({"error": "No exports folder found"}), 404
    
    # Get all Excel files
    excel_files = [f for f in os.listdir("exports") if f.endswith(".xlsx") and f != "combined_all_data.xlsx"]
    
    if not excel_files:
        return jsonify({"error": "No Excel files found"}), 404
    
    # Create combined workbook
    combined_wb = Workbook()
    combined_ws = combined_wb.active
    combined_ws.title = "All Data"
    
    # Create header row with styling
    headers = ["Sr No", "Name", "Address", "Phone", "Website", "Emails", "Exporter Type", "Date Added"]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = combined_ws.cell(1, col_num, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    combined_ws.column_dimensions['A'].width = 8
    combined_ws.column_dimensions['B'].width = 40
    combined_ws.column_dimensions['C'].width = 50
    combined_ws.column_dimensions['D'].width = 20
    combined_ws.column_dimensions['E'].width = 40
    combined_ws.column_dimensions['F'].width = 40
    combined_ws.column_dimensions['G'].width = 30
    combined_ws.column_dimensions['H'].width = 20
    
    current_row = 2
    sr_no = 1
    
    # Process each Excel file
    for excel_file in sorted(excel_files):
        filepath = os.path.join("exports", excel_file)
        
        # Extract exporter type from filename (remove .xlsx and replace underscores)
        exporter_type = excel_file.replace(".xlsx", "").replace("_", " ").title()
        
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Skip header row and read data
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:  # Skip empty rows
                    continue
                
                # Add data to combined sheet
                combined_ws.cell(current_row, 1, sr_no)
                combined_ws.cell(current_row, 2, row[1] if len(row) > 1 else "N/A")  # Name
                combined_ws.cell(current_row, 3, row[2] if len(row) > 2 else "N/A")  # Address
                combined_ws.cell(current_row, 4, row[3] if len(row) > 3 else "N/A")  # Phone
                combined_ws.cell(current_row, 5, row[4] if len(row) > 4 else "N/A")  # Website
                combined_ws.cell(current_row, 6, row[5] if len(row) > 5 else "N/A")  # Emails
                combined_ws.cell(current_row, 7, exporter_type)  # Exporter Type
                combined_ws.cell(current_row, 8, row[6] if len(row) > 6 else datetime.now().strftime("%Y-%m-%d %H:%M:%S"))  # Date
                
                # Apply alignment
                for col in range(1, 9):
                    combined_ws.cell(current_row, col).alignment = Alignment(vertical="top", wrap_text=True)
                
                sr_no += 1
                current_row += 1
            
            wb.close()
        except Exception as e:
            print(f"Error processing {excel_file}: {e}")
            continue
    
    # Save combined Excel file
    combined_filepath = os.path.join("exports", "combined_all_data.xlsx")
    combined_wb.save(combined_filepath)
    
    return jsonify({
        "success": True, 
        "filename": "combined_all_data.xlsx",
        "total_records": sr_no - 1
    })

@app.route("/", methods=["GET", "POST"])
def index():
    global current_results
    
    if request.method == "POST":
        query = request.form["query"]
        limit = int(request.form.get("limit", 10))
        
        # Clear previous results
        with current_results_lock:
            current_results = []
        
        # Start scraping in background thread
        def scrape_in_background():
            results = scrape_google_maps(query, limit)
            with current_results_lock:
                current_results.extend(results)
        
        thread = threading.Thread(target=scrape_in_background)
        thread.daemon = True
        thread.start()
        
        return render_template("index.html", streaming=True)

    return render_template("index.html", streaming=False)

if __name__ == "__main__":
    # Use threaded=False to prevent Flask from creating too many threads
    app.run(host="0.0.0.0", port=8580, debug=True, threaded=False)
